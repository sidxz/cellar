from __future__ import annotations
from pathlib import Path
from typing import AsyncIterator
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from cellar.application.export.renderers.base import RenderOptions
from cellar.application.export.renderers.excel_renderer import (
    ExcelRenderer,
    SPARKLINE_ROW_CAP,
)
from cellar.application.export.renderers.sparkline import (
    av_to_sparkline_snapshot as _av_to_sparkline_snapshot,
)
from cellar.application.export.row_streams.base import ColumnSpec, ExportRow


async def _batches(rows):
    for b in rows:
        yield b


@pytest.mark.asyncio
async def test_excel_writes_data_sheet_with_numeric_cells(tmp_path: Path):
    cols = [
        ColumnSpec(key="reg", header="Reg #", kind="text"),
        ColumnSpec(key="mw", header="MW", kind="number"),
        ColumnSpec(key="drc:rd1:ec:50.0::value", header="Mtb::EC50", kind="number", unit="µM"),
    ]
    rows = [[ExportRow(cells={"reg": "CV-1", "mw": 421.5, "drc:rd1:ec:50.0::value": 1.23})]]
    out = tmp_path / "out.xlsx"
    await ExcelRenderer().render(
        columns=cols, batches=_batches(rows), out_path=out,
        options=RenderOptions(), row_count_hint=1,
    )
    wb = load_workbook(out)
    ws = wb["Data"]
    assert ws["A1"].value == "Reg #"
    assert ws["B2"].value == 421.5
    assert isinstance(ws["B2"].value, float)
    assert ws["C2"].value == 1.23


def test_av_to_sparkline_snapshot_maps_fields():
    """_av_to_sparkline_snapshot must convert ActivityValue wire dict to the
    snapshot shape the sparkline renderer expects:
      - data_points with dose/response keys (remapped from x/y in raw_data)
      - fit with bottom/top/hill_slope/ec50 (ec50 comes from av['value'])
      - curve_class from curve_params
    """
    av = {
        "value": 1.5,  # the fitted EC50 — maps to fit["ec50"]
        "qualifier": None,
        "unit": "uM",
        "source": "dose_response",
        "raw_data": [
            {"x": 0.1, "y": 5.0, "is_excluded": False},
            {"x": 1.0, "y": 50.0},
            {"x": 10.0, "y": 95.0},
        ],
        "curve_params": {
            "hill_slope": 1.1,
            "top": 100.0,
            "bottom": 2.0,
            "num_points": 3,
            "curve_class": "active",
            "confidence_interval_low": None,
            "confidence_interval_high": None,
            "fit_quality_warnings": None,
        },
    }
    snap = _av_to_sparkline_snapshot(av)
    assert snap is not None
    assert snap["curve_class"] == "active"
    assert snap["fit"]["ec50"] == 1.5
    assert snap["fit"]["top"] == 100.0
    assert snap["fit"]["bottom"] == 2.0
    assert snap["fit"]["hill_slope"] == 1.1
    # x→dose, y→response; extra fields preserved
    assert snap["data_points"][0]["dose"] == 0.1
    assert snap["data_points"][0]["response"] == 5.0
    assert snap["data_points"][0]["is_excluded"] is False


def test_av_to_sparkline_snapshot_returns_none_without_fit():
    """_av_to_sparkline_snapshot must return None when curve_params lacks fit shape."""
    assert _av_to_sparkline_snapshot({}) is None
    assert _av_to_sparkline_snapshot(None) is None  # type: ignore[arg-type]
    assert _av_to_sparkline_snapshot({"curve_params": {"top": 100.0}}) is None


@pytest.mark.asyncio
async def test_excel_sparkline_uses_per_column_activity_lookup(tmp_path: Path):
    """The sparkline branch must look up the ActivityValue by the column's
    parent drc token, not by 'curve_snapshot'.

    Bug #2 regression: the old code did
        activity.get("curve_snapshot")
    which is always None because activity is {col_token: ActivityValue_dict}.
    The fix looks up activity[parent_token] and builds the snapshot from the
    ActivityValue's component fields.
    """
    rd_id = "rd1"
    parent_token = f"drc:{rd_id}"
    cols = [
        ColumnSpec(key="reg", header="Reg #", kind="text"),
        ColumnSpec(key=f"{parent_token}::plot", header="Mtb::Plot", kind="image_curve"),
    ]
    av_dict = {
        "value": 2.0,
        "qualifier": None,
        "unit": "uM",
        "source": "dose_response",
        "raw_data": [{"x": 1.0, "y": 50.0}],
        "curve_params": {
            "hill_slope": 1.0,
            "top": 100.0,
            "bottom": 0.0,
            "num_points": 1,
            "curve_class": "active",
            "confidence_interval_low": None,
            "confidence_interval_high": None,
            "fit_quality_warnings": None,
        },
    }
    rows = [[ExportRow(
        cells={"reg": "CV-1", f"{parent_token}::plot": ""},
        raw={"activity": {parent_token: av_dict}},
    )]]

    out = tmp_path / "out.xlsx"
    # Minimal valid 1×1 white PNG so openpyxl's XLImage can open it.
    import base64
    _1x1_png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChwGA60e6kgAAAABJRU5ErkJggg=="
    )
    with patch(
        "cellar.application.export.renderers.excel_renderer.render_sparkline_png",
        return_value=_1x1_png,
    ) as mock_render:
        await ExcelRenderer().render(
            columns=cols,
            batches=_batches(rows),
            out_path=out,
            options=RenderOptions(include_sparklines=True),
            row_count_hint=1,
        )

    # render_sparkline_png must have been called with the assembled snapshot,
    # NOT with None (which the old code always produced).
    mock_render.assert_called_once()
    called_snapshot = mock_render.call_args[0][0]
    assert called_snapshot is not None, (
        "Bug #2: render_sparkline_png was called with None — activity lookup "
        "used the wrong key (looked for 'curve_snapshot' on the activity dict "
        "instead of fetching the ActivityValue by parent token)"
    )
    assert called_snapshot.get("fit") is not None
    assert called_snapshot["fit"]["ec50"] == 2.0


@pytest.mark.asyncio
async def test_excel_notes_sheet_when_sparkline_cap_tripped(tmp_path: Path):
    cols = [ColumnSpec(key="reg", header="Reg #", kind="text")]
    big_rows = [[ExportRow(cells={"reg": f"CV-{i}"}) for i in range(SPARKLINE_ROW_CAP + 10)]]
    out = tmp_path / "out.xlsx"
    await ExcelRenderer().render(
        columns=cols, batches=_batches(big_rows), out_path=out,
        options=RenderOptions(), row_count_hint=SPARKLINE_ROW_CAP + 10,
    )
    wb = load_workbook(out)
    assert "Notes" in wb.sheetnames
    notes_text = "\n".join(str(c.value or "") for c in wb["Notes"]["A"])
    assert "Images omitted" in notes_text
