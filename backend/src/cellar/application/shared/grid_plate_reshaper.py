"""Reshape a plate-reader GRID file into the long ``Well, Value`` layout.

Plate readers (EnVision, PHERAstar, SpectraMax, …) export a matrix that mirrors
the plate — column numbers across the top, well-row letters down the side:

    Plate 1 — raw        <- optional title rows (skipped)
            1      2      3   ...
    A     2.31   5.10   1.92 ...
    B     4.20   6.33   2.07 ...

This module locates that grid block and flattens it to the ``Well, Value`` CSV
that ``ImportRunReadouts`` already consumes — so grid import reuses the entire
existing readout pipeline. Pure functions, no I/O beyond decoding the bytes.
"""

from __future__ import annotations

import csv
import io
import re

from cellar.domain.shared.errors import ValidationError

_ROW_LETTER_RE = re.compile(r"^[A-Z]{1,2}$")


def read_grid_cells(content: bytes, filename: str = "") -> list[list[str]]:
    """Decode CSV/TSV or XLSX bytes into a raw 2-D matrix of string cells."""
    is_xlsx = filename.lower().endswith((".xlsx", ".xlsm")) or content[:2] == b"PK"
    if is_xlsx:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            return [
                ["" if cell is None else str(cell) for cell in row]
                for row in ws.iter_rows(values_only=True)
            ]
        finally:
            wb.close()

    text = content.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]
    return [list(row) for row in csv.reader(io.StringIO(text), dialect)]


def _as_column_number(cell: str) -> int | None:
    """Return the positive integer a header cell represents, else None.

    Tolerates xlsx float headers (``"1.0"``) and ignores blanks/labels.
    """
    s = cell.strip()
    if not s:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    if f.is_integer() and f >= 1:
        return int(f)
    return None


def _row_letter(row: list[str]) -> str | None:
    """Return the well-row letter a data row starts with (A..AF), else None.

    The first non-empty cell must be a 1–2 char letter; otherwise this is a
    header/title row, not a data row.
    """
    for cell in row:
        c = cell.strip().upper()
        if not c:
            continue
        return c if _ROW_LETTER_RE.match(c) else None
    return None


def _locate_header_row(cells: list[list[str]]) -> tuple[int, dict[int, int]] | None:
    """Find the column-number header row.

    Returns ``(row_index, {cell_index: column_number})`` for the first row whose
    integer cells form ``1, 2, 3, …`` and that is immediately followed by a row
    starting with a well-row letter. Title rows above it are skipped naturally.
    """
    for idx, row in enumerate(cells):
        col_map = {ci: n for ci, cell in enumerate(row) if (n := _as_column_number(cell))}
        nums = sorted(col_map.values())
        looks_like_header = len(nums) >= 2 and nums == list(range(1, len(nums) + 1))
        if looks_like_header and idx + 1 < len(cells) and _row_letter(cells[idx + 1]):
            return idx, col_map
    return None


def reshape_grid_to_well_value_csv(content: bytes, filename: str = "") -> bytes:
    """Flatten a plate-reader grid file into a ``Well, Value`` CSV (bytes).

    Raises :class:`ValidationError` if no plate grid can be located or it holds
    no values.
    """
    cells = read_grid_cells(content, filename)
    located = _locate_header_row(cells)
    if located is None:
        raise ValidationError(
            "Could not find a plate grid: expected a header row of column numbers "
            "(1, 2, 3 …) above rows that start with a well-row letter (A, B, C …)."
        )

    header_idx, col_map = located
    out_rows: list[tuple[str, str]] = [("Well", "Value")]
    for row in cells[header_idx + 1 :]:
        letter = _row_letter(row)
        if letter is None:
            continue
        for ci, col_num in col_map.items():
            value = row[ci].strip() if ci < len(row) else ""
            if value:
                out_rows.append((f"{letter}{col_num}", value))

    if len(out_rows) == 1:
        raise ValidationError("The plate grid contained no readable values.")

    buf = io.StringIO()
    csv.writer(buf).writerows(out_rows)
    return buf.getvalue().encode("utf-8-sig")
