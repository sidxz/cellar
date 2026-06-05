from __future__ import annotations

import contextlib
from collections.abc import AsyncIterator
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cellar.application.export.renderers.base import RenderOptions
from cellar.application.export.renderers.sparkline import (
    SIZE_PRESETS,
    av_to_sparkline_snapshot,
    render_sparkline_png,
)
from cellar.application.export.renderers.structure import (
    STRUCTURE_SIZE_PRESETS,
    render_structure_pngs,
)
from cellar.application.export.row_streams.base import ColumnSpec, ExportRow

SPARKLINE_ROW_CAP = 5000


def _group_spans(columns: list[ColumnSpec]) -> list[tuple[int, int, str | None]]:
    """Compute consecutive-column spans by ColumnSpec.group.

    Returns a list of (start_index, end_index_inclusive, group_label).
    ``group_label`` is None for ungrouped columns.
    """
    spans: list[tuple[int, int, str | None]] = []
    if not columns:
        return spans
    start = 0
    cur = columns[0].group
    for i in range(1, len(columns)):
        if columns[i].group != cur:
            spans.append((start, i - 1, cur))
            start = i
            cur = columns[i].group
    spans.append((start, len(columns) - 1, cur))
    return spans


_GROUP_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
_GROUP_FONT = Font(bold=True, size=10, color="1E40AF")
_GROUP_ALIGN = Alignment(horizontal="center", vertical="center")
_GROUP_BORDER = Border(bottom=Side(style="medium", color="93C5FD"))


def _style_group_header(cell) -> None:
    cell.font = _GROUP_FONT
    cell.fill = _GROUP_FILL
    cell.alignment = _GROUP_ALIGN
    cell.border = _GROUP_BORDER


class ExcelRenderer:
    async def render(
        self,
        *,
        columns: list[ColumnSpec],
        batches: AsyncIterator[list[ExportRow]],
        out_path: Path,
        options: RenderOptions,
        row_count_hint: int,
    ) -> None:
        # NB: not using Workbook(write_only=True). Write-only sheets silently
        # ignore row_dimensions after `append` and don't reliably persist
        # column_dimensions when images are added — leaving structures + plots
        # squished into default-height rows. Memory budget at the 5K row cap
        # is comfortable (~150 MB peak) so the simplicity is worth it.
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Two-row header when any column carries a group label. Protocol
        # group labels (merged across each protocol's columns) on row 1,
        # per-column labels on row 2. With image_curve cells stacked under
        # the group, this mirrors the FE search grid 1:1.
        has_groups = any(c.group for c in columns)
        if has_groups:
            ws.append([""] * len(columns))  # placeholder for group row
            for start, end, label in _group_spans(columns):
                if not label:
                    continue
                if end > start:
                    ws.merge_cells(
                        start_row=1,
                        start_column=start + 1,
                        end_row=1,
                        end_column=end + 1,
                    )
                cell = ws.cell(row=1, column=start + 1, value=label)
                _style_group_header(cell)
            _HEADER_ROWS = 2
        else:
            _HEADER_ROWS = 1

        header_row = [(f"{c.header} ({c.unit})" if c.unit else c.header) for c in columns]
        ws.append(header_row)

        embed_images = options.include_sparklines and row_count_hint <= SPARKLINE_ROW_CAP
        sparkline_col_indices = [i for i, c in enumerate(columns) if c.kind == "image_curve"]
        structure_col_indices = [i for i, c in enumerate(columns) if c.kind == "image_structure"]
        size_preset = options.image_size if options.image_size in SIZE_PRESETS else "medium"
        sparkline_w, sparkline_h = SIZE_PRESETS[size_preset]
        struct_w, struct_h = STRUCTURE_SIZE_PRESETS.get(
            options.image_size, STRUCTURE_SIZE_PRESETS["medium"]
        )

        # Column widths: Excel uses character widths (~7 px each) plus a
        # ~5-px padding fudge. Empirically (width = px / 7 + 1) lands the
        # image inside the cell without horizontal clipping.
        if embed_images and sparkline_col_indices:
            for ci in sparkline_col_indices:
                ws.column_dimensions[get_column_letter(ci + 1)].width = sparkline_w / 7 + 1
        if embed_images and structure_col_indices:
            for ci in structure_col_indices:
                ws.column_dimensions[get_column_letter(ci + 1)].width = struct_w / 7 + 1

        # Per-row height in points (1 pt = 4/3 px). The tallest of the two
        # image kinds drives the row, plus a small margin so the image
        # doesn't kiss the gridline above.
        image_h_px = 0
        if embed_images and sparkline_col_indices:
            image_h_px = max(image_h_px, sparkline_h)
        if embed_images and structure_col_indices:
            image_h_px = max(image_h_px, struct_h)
        row_height_pt = (image_h_px * 0.75) + 6 if image_h_px else None

        tempfiles: list = []  # keep open until write
        rows_written = 0
        # Cache SMILES → tempfile path so repeated SMILES embed quickly.
        structure_cache: dict[str, str] = {}
        try:
            async for batch in batches:
                for row in batch:
                    out_row = []
                    for c in columns:
                        v = row.cells.get(c.key)
                        if c.kind == "number" and v is not None:
                            with contextlib.suppress(TypeError, ValueError):
                                v = float(v)
                        if c.kind in ("image_curve", "image_structure"):
                            out_row.append("")
                        else:
                            out_row.append(v)
                    ws.append(out_row)
                    rows_written += 1
                    excel_row = rows_written + _HEADER_ROWS  # +N for header rows
                    if row_height_pt is not None:
                        ws.row_dimensions[excel_row].height = row_height_pt

                    if embed_images and sparkline_col_indices:
                        # Each image_curve column is keyed by "drc:<rd_id>::plot".
                        # The parent activity token is "drc:<rd_id>" — one
                        # ActivityValue (serialised via dataclasses.asdict) per
                        # readout-def.  ActivityValue has no pre-built
                        # curve_snapshot field; we assemble a snapshot-compatible
                        # dict from the component fields that the sparkline
                        # renderer expects.
                        activity = (row.raw.get("activity") or {}) if row.raw else {}
                        for ci in sparkline_col_indices:
                            col = columns[ci]
                            if "::" not in col.key:
                                continue
                            parent_token = col.key.rsplit("::", 1)[0]
                            av = activity.get(parent_token)
                            if not av:
                                continue
                            snapshot = av_to_sparkline_snapshot(av)
                            png = (
                                render_sparkline_png(snapshot, size=size_preset)
                                if snapshot
                                else None
                            )
                            if not png:
                                continue
                            with NamedTemporaryFile(suffix=".png", delete=False) as tf:
                                tf.write(png)
                            tempfiles.append(tf.name)
                            img = XLImage(tf.name)
                            img.width = sparkline_w
                            img.height = sparkline_h
                            cell_ref = f"{get_column_letter(ci + 1)}{excel_row}"
                            ws.add_image(img, cell_ref)

                    if embed_images and structure_col_indices:
                        smiles = (row.raw or {}).get("smiles")
                        if not smiles:
                            continue
                        tf_path = structure_cache.get(smiles)
                        if tf_path is None:
                            pngs = render_structure_pngs([smiles], size=options.image_size)
                            png = pngs.get(smiles)
                            if not png:
                                continue
                            with NamedTemporaryFile(suffix=".png", delete=False) as tf:
                                tf.write(png)
                            tempfiles.append(tf.name)
                            tf_path = tf.name
                            structure_cache[smiles] = tf_path
                        for ci in structure_col_indices:
                            img = XLImage(tf_path)
                            img.width = struct_w
                            img.height = struct_h
                            cell_ref = f"{get_column_letter(ci + 1)}{excel_row}"
                            ws.add_image(img, cell_ref)

            notes = wb.create_sheet("Notes")
            notes.append([options.title])
            notes.append([f"Rows: {rows_written}"])
            if not embed_images and row_count_hint > SPARKLINE_ROW_CAP:
                notes.append(
                    [
                        f"Images omitted: row count {row_count_hint} "
                        f"exceeds {SPARKLINE_ROW_CAP} cap."
                    ]
                )

            wb.save(out_path)
        finally:
            import os

            for p in tempfiles:
                with contextlib.suppress(OSError):
                    os.unlink(p)
