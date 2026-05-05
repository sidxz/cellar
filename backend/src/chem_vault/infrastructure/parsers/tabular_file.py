"""Shared tabular file parser — xlsx + csv into a unified record stream.

Used by importers that consume row-based data (plate maps, readout files,
long-format run files). Format is detected from the filename extension and
the file's magic bytes (xlsx is a zip; csv is plain text).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterator
from dataclasses import dataclass


class TabularParseError(ValueError):
    """Raised when a file cannot be parsed as a supported tabular format."""


@dataclass
class ParsedTable:
    """An immutable, in-memory tabular file."""

    headers: list[str]
    rows: list[dict[str, str]]
    source_format: str  # "xlsx" | "csv"

    def iter_rows(self) -> Iterator[dict[str, str]]:
        return iter(self.rows)

    @property
    def row_count(self) -> int:
        return len(self.rows)


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

_XLSX_MAGIC = b"PK\x03\x04"  # ZIP signature; xlsx is a zip container


def _detect_format(content: bytes, filename: str) -> str:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or content[:4] == _XLSX_MAGIC:
        return "xlsx"
    return "csv"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def parse_tabular(content: bytes, filename: str = "") -> ParsedTable:
    """Parse a tabular file into a ParsedTable.

    Detects xlsx vs csv from extension + magic bytes. xlsx is read with
    openpyxl in read-only mode; csv falls back to delimiter-sniffing
    against `,;\t|` and tolerates UTF-8 BOM.

    Raises TabularParseError on unparseable input.
    """
    if not content:
        raise TabularParseError("File is empty")

    fmt = _detect_format(content, filename)
    if fmt == "xlsx":
        return _parse_xlsx(content)
    return _parse_csv(content)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _parse_csv(content: bytes) -> ParsedTable:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1", errors="replace")

    if not text.strip():
        raise TabularParseError("CSV file is empty")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = list(reader)
    if not all_rows:
        raise TabularParseError("CSV has no rows")

    headers = [(h or "").strip() for h in all_rows[0]]
    if not any(headers):
        raise TabularParseError("CSV has no header row")

    rows: list[dict[str, str]] = []
    for raw in all_rows[1:]:
        if not any((c or "").strip() for c in raw):
            continue
        row = {
            headers[i]: (raw[i].strip() if i < len(raw) and raw[i] is not None else "")
            for i in range(len(headers))
        }
        rows.append(row)

    return ParsedTable(headers=headers, rows=rows, source_format="csv")


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


def _parse_xlsx(content: bytes) -> ParsedTable:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dep is declared
        raise TabularParseError("openpyxl is required to parse xlsx") from exc

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise TabularParseError(f"Cannot read xlsx: {exc}") from exc

    try:
        ws = wb.active
        if ws is None:
            raise TabularParseError("xlsx has no active sheet")

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise TabularParseError("xlsx has no header row") from exc

        headers = [_cell_to_str(c).strip() for c in header_row]
        # Trim trailing empty header columns
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            raise TabularParseError("xlsx has no header row")

        rows: list[dict[str, str]] = []
        for raw in rows_iter:
            cells = [_cell_to_str(c) for c in raw]
            if not any(c.strip() for c in cells):
                continue
            row = {
                headers[i]: (cells[i].strip() if i < len(cells) else "")
                for i in range(len(headers))
            }
            rows.append(row)
    finally:
        wb.close()

    return ParsedTable(headers=headers, rows=rows, source_format="xlsx")


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value) if abs(value) < 1e-4 or abs(value) >= 1e16 else f"{value}"
    return str(value)
